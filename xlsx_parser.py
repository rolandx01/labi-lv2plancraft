"""
xlsx_parser.py - Liest LV-XLSX-Dateien und gibt Positionen zurück.

Unterstützt zwei Eingabe-Formate:

FORMAT A — Plancraft-kompatibel (5 Spalten):
    Menge | Einheit | Kurztext | Langtext | Einheitspreis (€)
    (typischerweise von Plancraft-Export oder meinem Tool generiert)
    -> Nur Einheiten-Normalisierung nötig

FORMAT B — Förderantrag-LV (9 Spalten):
    Pos. | Menge | Einheit | Beschreibung | Preis | Gesamt | ISFP | Förderbetrag | ...
    (typischerweise von BAFA/KfW-Antragssoftware)
    -> Spalten-Mapping nötig

Wichtig: KEINE KI, nur deterministisches Parsen.
"""

from dataclasses import dataclass, field
from typing import List, Optional
import re
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from lv_parser import Position, normalisiere_einheit


@dataclass
class XlsxResult:
    """Ergebnis eines XLSX-Pars."""
    positionen: List[Position] = field(default_factory=list)
    fehler: List[str] = field(default_factory=list)
    warnungen: List[str] = field(default_factory=list)
    erkanntes_format: str = ""  # "plancraft" oder "foerderantrag" oder "unbekannt"
    seiten_anzahl: int = 1  # XLSX hat keine Seiten, hier = Anzahl Datenzeilen


def _parse_plancraft_format(ws: Worksheet) -> List[Position]:
    """Liest ein XLSX im 5-Spalten-Plancraft-Format.

    Sektion-Header (Zeile mit Text in Spalte C, aber None in Spalte A/B/E)
    werden als Pseudo-Positionen behalten (Menge=1, Einheit="psch."), damit
    die Sektion-Struktur in Plancraft sichtbar bleibt.
    """
    positionen = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Header überspringen
        # row = (menge, einheit, kurztext, langtext, einheitspreis)
        if not row:
            continue

        # Padding falls Zeile kürzer als 5 Spalten
        menge_raw, einheit_raw, kurztext, langtext, einheitspreis = (list(row) + [None]*5)[:5]

        # Sektion-Header erkennen: Spalte C hat Text, A+B+E leer
        # Bsp: (None, None, "Allgemein", None, None) oder (None, None, "Stundenlohnarbeiten", None, None)
        ist_sektion_header = (
            (menge_raw is None or str(menge_raw).strip() == "")
            and (einheit_raw is None or str(einheit_raw).strip() == "")
            and kurztext is not None and str(kurztext).strip() != ""
            and (langtext is None or str(langtext).strip() == "")
            and (einheitspreis is None or str(einheitspreis).strip() == "")
        )
        if ist_sektion_header:
            # Als Pseudo-Position: Menge=1, Einheit=psch., Preis leer
            # → erscheint in Plancraft als "1 psch. Sektionsname" (Plancraft akzeptiert das)
            positionen.append(Position(
                menge=1,
                einheit="psch.",
                kurztext=str(kurztext).strip(),
                langtext="",
                einheitspreis=None,
                pos_nr="",  # Keine Positionsnummer für Sektion
                seite=1,
            ))
            continue

        # Komplett leere Zeile überspringen
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        # Menge
        try:
            menge = parse_zahl_smart(menge_raw)
        except (ValueError, TypeError):
            menge = None

        # Einheit normalisieren
        einheit = normalisiere_einheit(str(einheit_raw)) if einheit_raw else None

        # Einheitspreis
        try:
            ep = parse_zahl_smart(einheitspreis)
        except (ValueError, TypeError):
            ep = None

        positionen.append(Position(
            menge=menge,
            einheit=einheit,
            kurztext=str(kurztext).strip() if kurztext else "",
            langtext=str(langtext).strip() if langtext else "",
            einheitspreis=ep,
            seite=1,
        ))

    return positionen


def parse_deutsche_zahl(s: str) -> Optional[float]:
    """Parst eine Zahl, die im deutschen ODER englischen Format vorliegen kann.

    Beispiele:
        "2.433,60"  → 2433.60  (deutsch: Punkt=Tausender, Komma=Dezimal)
        "1.000"     → 1000     (deutsch: nur Tausender)
        "2433.60"   → 2433.60  (englisch: Punkt=Dezimal)
        "1,000"     → 1000     (englisch: nur Tausender)
        "1.5"       → 1.5      (englisch: Dezimal)
        "1,5"       → 1.5      (deutsch: Dezimal)
        "0,5"       → 0.5
        ""          → None
        "abc"       → None
    """
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None

    # Entferne Tausendertrennzeichen und normalisiere Dezimaltrennzeichen
    # Strategie: Wenn BEIDE (Komma UND Punkt) vorkommen, ist das letzte das Dezimaltrennzeichen
    # Wenn nur eines vorkommt:
    #   - Wenn 3 Stellen nach dem Trenner → Tausender (bei Komma) ODER Dezimal (bei Punkt)
    #   - Sonst → Dezimal
    hat_komma = "," in s
    hat_punkt = "." in s

    if hat_komma and hat_punkt:
        # Beide vorhanden: das LETZTE Trennzeichen ist das Dezimaltrennzeichen
        letzte_komma = s.rfind(",")
        letzte_punkt = s.rfind(".")
        if letzte_komma > letzte_punkt:
            # Komma ist Dezimal → "2,433.60" oder "2.433,60" → Punkt=Tausender, Komma=Dezimal
            s = s.replace(".", "").replace(",", ".")
        else:
            # Punkt ist Dezimal → "2,433.60" → Komma=Tausender, Punkt=Dezimal
            s = s.replace(",", "")
    elif hat_komma:
        # Nur Komma: Wenn nach Komma genau 3 Stellen UND keine führenden 0 → Tausender
        teile = s.split(",")
        if len(teile) == 2 and len(teile[1]) == 3 and teile[0].isdigit():
            s = s.replace(",", "")  # "1,000" → "1000"
        else:
            s = s.replace(",", ".")  # "1,5" → "1.5"
    elif hat_punkt:
        # Nur Punkt: Wenn nach Punkt genau 3 Stellen UND keine führenden 0 → Tausender
        teile = s.split(".")
        if len(teile) == 2 and len(teile[1]) == 3 and teile[0].isdigit():
            s = s.replace(".", "")  # "1.000" → "1000"
        else:
            # "1.5" → bleibt "1.5" (Dezimal)
            pass

    try:
        return float(s)
    except ValueError:
        return None


# --- Anwendungs-Spezialfall: Excel-werte (z.B. 3933.6000000000004) ---
# Excel gibt oft Floats mit Rundungsfehlern zurück. Die wollen wir nicht als "Zahl mit Komma" interpretieren.
# Wenn der String als Python-Float parsed werden kann UND kein Komma enthält, ist es ein "reiner" Float.
def parse_zahl_smart(wert) -> Optional[float]:
    """Universeller Zahlen-Parser: akzeptiert int, float, str."""
    if wert is None:
        return None
    if isinstance(wert, (int, float)):
        return float(wert)
    if isinstance(wert, str):
        return parse_deutsche_zahl(wert)
    return None


def _parse_foerderantrag_format(ws: Worksheet) -> List[Position]:
    """Liest ein XLSX im Förderantrag-LV-Format (9 Spalten).

    Erwartete Spalten (A-I):
        A: Pos. (z.B. "1. ", "2. ", oder leer bei Sub-Position)
        B: Menge
        C: Einheit
        D: Beschreibung (Langtext)
        E: Preis (= Einheitspreis)
        F: Gesamt
        G: ISFP
        H: Förderbetrag
        I: (optional)

    Wichtig: Manche LVs (z.B. BAFA-Anträge) haben Sub-Positionen ohne eigene
    Pos-Nr in Spalte A. Diese erben den Kontext der vorherigen Hauptposition
    und werden als separate Positionen mit eigenem Kurztext ausgegeben.
    """
    positionen = []

    # Header-Reihe finden (typischerweise Reihe 4, aber flexibel)
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] and "pos" in str(row[0]).lower():
            header_row_idx = i
            break

    if header_row_idx is None:
        return positionen

    # Tracken: aktuelle Hauptposition (für Sub-Positionen)
    aktuelle_hauptposition = None  # tuple (pos_nr, kurztext_haupt)
    letzte_pos_war_sektion_header = False  # Damit wir nicht in falsche Richtung tracken

    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        if not row or len(row) < 4:
            continue

        pos_nr_raw, menge_raw, einheit_raw, beschreibung, preis_raw, gesamt_raw, *_ = row

        # Strings für Vergleiche
        pos_nr_str = str(pos_nr_raw).strip() if pos_nr_raw is not None else ""
        beschreibung_str = str(beschreibung).strip() if beschreibung is not None else ""
        einheit_str = str(einheit_raw).strip() if einheit_raw is not None else ""
        preis_str = str(preis_raw).strip() if preis_raw is not None else ""
        gesamt_str = str(gesamt_raw).strip() if gesamt_raw is not None else ""

        # Komplett leere Zeile überspringen
        if (not pos_nr_str or pos_nr_str in ("None", "·", ".")) \
                and not beschreibung_str and not einheit_str \
                and not preis_str and not gesamt_str:
            continue

        # Reine Footer/Summen-Zeilen erkennen (z.B. "Gesamt:", "MWST", "Förderung")
        # Diese haben KEINE pos_nr, KEIN einheit, ABER text in D oder E
        footer_keywords = ("gesamtkosten", "gesamt:", "mwst", "förderung", "nach abzug",
                           "eigenleistung", "eigenmittel", "bank kfw", "verkauf", "solar")
        ist_footer_zeile = (
            not pos_nr_str
            and not einheit_str
            and not preis_str
            and (beschreibung_str.lower().startswith(footer_keywords)
                 or gesamt_str.lower().startswith(footer_keywords))
        )
        if ist_footer_zeile:
            continue

        # Sektion-Trenner: Zeile mit Beschreibung in Spalte A oder D, aber keine Pos-Nr/Menge/Einheit/Preis
        # Beispiele: "Dämmung des Daches" (in A), "Fliesenarbeiten Bäder" (in D)
        # WICHTIG: Eine Sektion hat NICHTS in den Datenspalten (B/C/E).
        # Wenn B/C/E befüllt sind, ist es eine echte Position (ggf. Sub-Position).
        hat_daten = (
            (menge_raw is not None and str(menge_raw).strip() not in ("", "None", "·", "."))
            or einheit_str
            or preis_str
        )
        # Sektion-Verdacht: nur D hat Text, A ist leer, B/C/E sind leer
        # ABER: Wenn die VORHERIGE Zeile eine Hauptposition war, ist das vermutlich
        # eine Sub-Position (z.B. R13 folgt auf R12 mit "1. " = Hauptposition)
        ist_vermutlich_sub = (
            not pos_nr_str
            and not hat_daten
            and beschreibung_str
            and aktuelle_hauptposition is not None
        )
        ist_sektion_header = (
            (not pos_nr_str or pos_nr_str in ("None", "·", "."))
            and not hat_daten  # Sektion hat NIE Daten in B/C/E
            and beschreibung_str
            and not ist_vermutlich_sub  # Nur Sektion wenn keine vorige Hauptposition
        )
        # Spezialfall: Sektion in Spalte A (z.B. "Zimmer Innentüren")
        if pos_nr_str and not re.match(r"^\d+\.?\s*$", pos_nr_str) and not hat_daten:
            ist_sektion_header = True

        if ist_sektion_header:
            # FIX: Sektion-Header als Pseudo-Position behalten
            s_name = pos_nr_str if pos_nr_str and pos_nr_str not in ("None", "·", ".") else beschreibung_str
            positionen.append(Position(
                menge=1,
                einheit="psch.",
                pos_nr="",
                kurztext=s_name[:120],
                langtext="",
                einheitspreis=None,
                seite=1,
            ))
            aktuelle_hauptposition = None  # Sektion-Header beendet vorherigen Block
            continue

        # Echte Position: pos_nr ist "1." oder "1" oder leer (Sub-Position)
        pos_nr = ""
        if pos_nr_raw and re.match(r"^\d+\.?\s*$", pos_nr_str):
            m = re.match(r"^(\d+)", pos_nr_str)
            if m:
                pos_nr = m.group(1)
                # Das ist eine Hauptposition — tracken für Sub-Items
                # Kurztext = erste 120 Zeichen der Beschreibung
                aktuelle_hauptposition = (pos_nr, beschreibung_str[:80] if beschreibung_str else "")
        elif not pos_nr_str and aktuelle_hauptposition and beschreibung_str:
            # Sub-Position: erbt pos_nr der Hauptposition mit Suffix
            pos_nr = f"{aktuelle_hauptposition[0]}.sub"
        elif not pos_nr_str and aktuelle_hauptposition and not beschreibung_str:
            # Hat weder pos_nr noch Beschreibung — Müll
            continue

        # Menge
        try:
            menge = parse_zahl_smart(menge_raw)
        except (ValueError, TypeError):
            menge = None

        # Einheit
        einheit = normalisiere_einheit(einheit_str) if einheit_str else None

        # Einheitspreis (Spalte E)
        try:
            ep = parse_zahl_smart(preis_raw)
        except (ValueError, TypeError):
            ep = None

        # Kurztext: erste Zeile der Beschreibung (max 120 Zeichen)
        kurztext = beschreibung_str[:120] if beschreibung_str else ""

        # Wenn weder Kurztext noch Langtext, sondern nur ein Preis — versuch
        # einen sinnvollen Kurztext zu generieren aus Kontext
        if not kurztext and aktuelle_hauptposition:
            kurztext = f"{aktuelle_hauptposition[1]} (Sub-Position)"[:120]

        # Langtext = ganze Beschreibung
        langtext = beschreibung_str

        if not kurztext and not langtext:
            continue

        positionen.append(Position(
            menge=menge,
            einheit=einheit,
            pos_nr=pos_nr,
            kurztext=kurztext,
            langtext=langtext,
            einheitspreis=ep,
            seite=1,
        ))

    return positionen


def _parse_kostenschaetzung_format(ws: Worksheet) -> List[Position]:
    """Liest ein XLSX im Kostenschätzungs-Format (Architekten-/Bauplaner-Listen).

    Erwartete Spalten (typisch):
        A: Pos. (Zahl oder leer bei Sektion)
        B: Gewerk (= Titel/Kurztext)
        C: Einheit
        D: Menge
        E: Preis/Stück (= Einheitspreis)
        F: Summe (netto) — oft befüllt auch wenn E leer
        G: Annahmen / Notizen (= Langtext)
        H-L: oft leer

    Sektion-Trenner (z.B. "Fenster und Eingangstüren", "Heizung"):
        Spalte A = Text, Spalte C leer, keine Zahlen → überspringen

    Summen-Zeilen (z.B. "Gesamtsumme Gewerk Fenster"):
        Spalte A leer, Spalte B leer, nur Spalte F hat Zahl → überspringen

    Echte Position:
        Spalte A = Zahl (Pos) ODER
        Spalte B hat Text UND Spalte F hat Zahl
    """
    positionen = []

    # Header-Reihe finden: Zeile wo A="Pos." und C="Einheit"
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if (row and len(row) >= 5
                and row[0] and "pos" in str(row[0]).lower()
                and row[2] and "einheit" in str(row[2]).lower()):
            header_row_idx = i
            break

    if header_row_idx is None:
        return positionen

    import re

    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), header_row_idx + 1):
        if not row or len(row) < 5:
            continue

        pos_nr_raw, gewerk, einheit_raw, menge_raw, preis_raw, summe_raw, annahmen, *rest = row

        # Sektion-Trenner: A=Text (keine Zahl), C leer, keine echte Daten
        pos_nr_str = str(pos_nr_raw).strip() if pos_nr_raw is not None else ""
        gewerk_str = str(gewerk).strip() if gewerk is not None else ""

        # Wenn Spalte A ein Wort ist (Sektion-Header wie "Fenster und Eingangstüren")
        if pos_nr_str and not re.match(r"^\d+", pos_nr_str):
            # Sektion-Trenner
            if not einheit_raw and not menge_raw and not preis_raw:
                # FIX: Sektion-Header als Pseudo-Position behalten (Menge=1, Einheit=psch.)
                # Begründung: Plancraft braucht die Sektion-Struktur, sonst weiss Labi nicht
                # welche Position zu welcher Sektion gehört. Sektion in Kurztext.
                positionen.append(Position(
                    menge=1,
                    einheit="psch.",
                    pos_nr="",
                    kurztext=pos_nr_str,  # Sektion-Name als Kurztext
                    langtext="",
                    einheitspreis=None,
                    seite=1,
                ))
                continue
            # Andernfalls ist es eine echte Position mit Text in Spalte A (selten)

        # Reine Summen-Zeile: A leer, B leer, nur F hat Zahl
        if (not pos_nr_str or pos_nr_str in ("None", "·", ".")) \
                and (not gewerk_str or gewerk_str in ("None", "·", ".")) \
                and summe_raw is not None and einheit_raw is None:
            continue

        # Leere Zeile komplett überspringen
        if (not pos_nr_str or pos_nr_str in ("None", "·", ".")) \
                and (not gewerk_str or gewerk_str in ("None", "·", ".")) \
                and menge_raw is None and preis_raw is None and summe_raw is None:
            continue

        # Menge parsen
        try:
            menge = parse_zahl_smart(menge_raw)
        except (ValueError, TypeError):
            menge = None

        # Einheit normalisieren
        einheit = normalisiere_einheit(str(einheit_raw)) if einheit_raw else None

        # Preis parsen (E = Preis/Stück)
        try:
            ep = parse_zahl_smart(preis_raw)
        except (ValueError, TypeError):
            ep = None

        # Positionsnummer: nur die Zahl extrahieren
        pos_nr = ""
        if pos_nr_raw:
            m = re.match(r"^(\d+)", pos_nr_str)
            if m:
                pos_nr = m.group(1)

        # Kurztext = Gewerk (Titel)
        # Langtext = Annahmen/Notizen (Detail)
        kurztext = gewerk_str
        langtext = str(annahmen).strip() if annahmen else ""

        # Wenn beides fehlt, ist die Zeile doch Müll
        if not kurztext and not langtext:
            continue

        positionen.append(Position(
            menge=menge,
            einheit=einheit,
            pos_nr=pos_nr,
            kurztext=kurztext,
            langtext=langtext,
            einheitspreis=ep,
            seite=1,
        ))

    return positionen


def erkenne_format(ws: Worksheet) -> str:
    """Erkennt welches XLSX-Format vorliegt.

    Returns:
        "plancraft" (5 Spalten, "Menge" in A1)
        "foerderantrag" (9 Spalten, "Pos." oder "Beschreibung" in Header)
        "kostenschaetzung" (Architekten-LV mit "Pos." + "Einheit" + "Preis/Stück")
        "unbekannt"
    """
    # Sammle alle Zeilen in den ersten 10 (manche Formate haben Titel in Zeile 1, Header in 3)
    alle_header_zeilen = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    if not alle_header_zeilen:
        return "unbekannt"

    # Plancraft-Format: 5 Spalten, "Menge" in Spalte A (Zeile 1)
    if len(alle_header_zeilen[0]) == 5 and alle_header_zeilen[0][0] and "menge" in str(alle_header_zeilen[0][0]).lower():
        return "plancraft"

    # Förderantrag-Format: 9+ Spalten, "Beschreibung" in Header (irgendwo in Top 5)
    for row in alle_header_zeilen[:5]:
        if row and any(c and "beschreibung" in str(c).lower() for c in row if c):
            return "foerderantrag"

    # Kostenschätzungs-Format: "Pos." in Spalte A + "Einheit" in Spalte C
    # (in irgendeiner der ersten 10 Zeilen)
    for row in alle_header_zeilen:
        if (row and len(row) >= 3
                and row[0] and "pos" in str(row[0]).lower()
                and row[2] and "einheit" in str(row[2]).lower()):
            return "kostenschaetzung"

    # Fallback: Förderantrag mit "Pos." in Spalte A (aber kein "Beschreibung" in Top 5)
    for row in alle_header_zeilen:
        if row and len(row) >= 5 and row[0] and "pos" in str(row[0]).lower():
            return "foerderantrag"

    return "unbekannt"


def parse_xlsx(xlsx_pfad: str) -> XlsxResult:
    """Parst eine XLSX-LV-Datei und gibt Positionen zurück.

    Args:
        xlsx_pfad: Absoluter Pfad zur XLSX-Datei

    Returns:
        XlsxResult mit positionen, fehler, warnungen
    """
    result = XlsxResult()

    try:
        wb = load_workbook(xlsx_pfad, data_only=True)
    except Exception as e:
        result.fehler.append(f"XLSX konnte nicht gelesen werden: {e}")
        return result

    # Erstes Sheet verwenden
    if not wb.sheetnames:
        result.fehler.append("XLSX hat keine Sheets.")
        return result

    ws = wb[wb.sheetnames[0]]
    result.erkanntes_format = erkenne_format(ws)
    result.seiten_anzahl = ws.max_row

    if result.erkanntes_format == "plancraft":
        result.positionen = _parse_plancraft_format(ws)
    elif result.erkanntes_format == "foerderantrag":
        result.positionen = _parse_foerderantrag_format(ws)
    elif result.erkanntes_format == "kostenschaetzung":
        result.positionen = _parse_kostenschaetzung_format(ws)
    else:
        result.fehler.append(
            f"Unbekanntes XLSX-Format. Erste Zeile: "
            f"{[str(c)[:30] if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]}"
        )
        return result

    # Validierung
    if not result.positionen:
        result.fehler.append("XLSX-Parser hat keine Positionen gefunden.")
    else:
        # Sammle unbekannte Einheiten (z.B. StWo im Langtext)
        unbekannte_einheiten = set()
        for p in result.positionen:
            if p.langtext:
                import re
                matches = re.findall(r"\d+(?:[.,]\d+)?\s+([A-Za-z]+Wo)\b", p.langtext)
                for m in matches:
                    unbekannte_einheiten.add(m)

        if unbekannte_einheiten:
            result.warnungen.append(
                f"{len(unbekannte_einheiten)} zusammengesetzte Einheit(en) gefunden "
                f"({', '.join(sorted(unbekannte_einheiten))}), "
                "die Plancraft NICHT direkt unterstützt."
            )

        # Warnung, wenn viele Positionen ohne Menge
        ohne_menge = sum(1 for p in result.positionen if p.menge is None)
        if ohne_menge > len(result.positionen) * 0.2:
            result.warnungen.append(
                f"{ohne_menge} von {len(result.positionen)} Positionen haben keine erkannte Menge."
            )

    return result


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python xlsx_parser.py <xlsx-pfad>")
        sys.exit(1)

    ergebnis = parse_xlsx(sys.argv[1])
    print(f"Format: {ergebnis.erkanntes_format}")
    print(f"Positionen: {len(ergebnis.positionen)}")
    print(f"Fehler: {ergebnis.fehler}")
    print(f"Warnungen: {ergebnis.warnungen}")
    print()
    for i, p in enumerate(ergebnis.positionen[:5], 1):
        print(f"  [{i}] Pos {p.pos_nr}: {p.kurztext[:50]}")
        print(f"      Menge: {p.menge} {p.einheit} | EP: {p.einheitspreis}")
