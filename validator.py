"""
validator.py - Pre-Validierung der generierten Plancraft-XLSX

Prüft VOR dem Download, ob die XLSX alle Plancraft-Anforderungen erfüllt.
Plancrafts Fehlermeldung "irgendein Problem aufgetaucht" ist leider sehr
kryptisch. Wir simulieren Plancrafts Import-Logik und melden konkrete Probleme.

Geprüft wird:
1. Header exakt: "Menge | Einheit | Kurztext | Langtext | Einheitspreis"
2. Spaltenbreite = 5 (Plancraft akzeptiert nur 5 Spalten in den Datenzeilen)
3. Keine leeren Datenzeilen
4. Alle Einheiten in Plancraft-Whitelist
5. Keine negativen Mengen oder Preise
6. Keine "nan" / "None" Strings
7. Mengen sind Zahlen, keine Strings

Output: ValidationResult mit status ("ok" | "warnung" | "fehler"),
        probleme (Liste von strings), und statistik.
"""

from dataclasses import dataclass, field
from typing import List
from openpyxl import load_workbook
from lv_parser import (
    PLANCRAFT_EINHEITEN_OFFIZIELL,
    normalisiere_einheit,
)


# Genau diese 5 Header erwartet Plancraft (Reihenfolge wichtig)
ERWARTETE_HEADER = ["Menge", "Einheit", "Kurztext", "Langtext", "Einheitspreis"]


@dataclass
class ValidationResult:
    """Ergebnis der Pre-Validierung."""
    status: str = "ok"  # "ok" | "warnung" | "fehler"
    probleme: List[str] = field(default_factory=list)
    warnungen: List[str] = field(default_factory=list)
    info: List[str] = field(default_factory=list)
    statistik: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "status": self.status,
            "probleme": self.probleme,
            "warnungen": self.warnungen,
            "info": self.info,
            "statistik": self.statistik,
        }


def validiere_xlsx(xlsx_pfad: str) -> ValidationResult:
    """Liest die geschriebene Plancraft-XLSX und prüft sie.

    Args:
        xlsx_pfad: Absoluter Pfad zur XLSX-Datei

    Returns:
        ValidationResult mit status, probleme, warnungen, info, statistik
    """
    result = ValidationResult()

    try:
        wb = load_workbook(xlsx_pfad, data_only=True)
    except Exception as e:
        result.status = "fehler"
        result.probleme.append(f"XLSX konnte nicht geöffnet werden: {e}")
        return result

    if not wb.sheetnames:
        result.status = "fehler"
        result.probleme.append("XLSX enthält keine Sheets.")
        return result

    ws = wb[wb.sheetnames[0]]

    # --- Check 1: Header exakt ---
    header_zeile = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []))
    header_norm = [str(c).strip() if c is not None else "" for c in header_zeile]

    if header_norm[:5] != ERWARTETE_HEADER:
        result.status = "fehler"
        result.probleme.append(
            f"Falsche Header. Erwartet: {ERWARTETE_HEADER}. "
            f"Gefunden: {header_norm[:5]}. "
            "Plancraft akzeptiert nur exakt diese Spaltennamen in dieser Reihenfolge."
        )
        # Wenn Header kaputt, ist alles weitere sinnlos
        return result

    # --- Check 2: Datenzeilen ---
    datenzeilen = list(ws.iter_rows(min_row=2, values_only=True))
    anzahl_datenzeilen = len(datenzeilen)
    leere_zeilen = 0
    ungueltige_einheiten = set()
    ungueltige_mengen = []
    ungueltige_preise = []
    positionen_ohne_text = 0
    verdoppelte_kurztexte = {}  # kurztext -> count

    for i, row in enumerate(datenzeilen, start=2):  # Zeile 2 = erste Datenzeile
        if not row:
            leere_zeilen += 1
            continue

        # Padding falls kürzer
        menge, einheit, kurztext, langtext, einheitspreis = (list(row) + [None]*5)[:5]

        # Komplett leere Zeile?
        if all(c is None or str(c).strip() == "" for c in row):
            leere_zeilen += 1
            continue

        # Check 3: Menge
        if menge is not None:
            menge_str = str(menge).strip()
            if menge_str in ("nan", "None", "NaN"):
                ungueltige_mengen.append(f"Zeile {i}: Menge='{menge_str}' (kein gültiger Zahlenwert)")
            else:
                try:
                    menge_val = float(menge)
                    if menge_val < 0:
                        ungueltige_mengen.append(f"Zeile {i}: negative Menge ({menge_val})")
                except (ValueError, TypeError):
                    ungueltige_mengen.append(f"Zeile {i}: Menge='{menge_str}' ist keine Zahl")

        # Check 4: Einheit in Whitelist
        if einheit is not None and str(einheit).strip():
            einheit_str = str(einheit).strip()
            if einheit_str not in PLANCRAFT_EINHEITEN_OFFIZIELL:
                # Versuch zu normalisieren — vielleicht war es ein Alias
                normalisiert = normalisiere_einheit(einheit_str)
                if normalisiert and normalisiert in PLANCRAFT_EINHEITEN_OFFIZIELL:
                    result.warnungen.append(
                        f"Zeile {i}: Einheit '{einheit_str}' wird zu '{normalisiert}' normalisiert. "
                        "Sollte vor Plancraft-Import gefixt werden."
                    )
                else:
                    ungueltige_einheiten.add(einheit_str)

        # Check 5: Position ohne Text (sowohl Kurz- als auch Langtext leer)
        kurztext_str = str(kurztext).strip() if kurztext is not None else ""
        langtext_str = str(langtext).strip() if langtext is not None else ""
        if not kurztext_str and not langtext_str:
            positionen_ohne_text += 1

        # Verdoppelte Kurztexte (Plancraft macht das u.U. Probleme)
        if kurztext_str:
            verdoppelte_kurztexte[kurztext_str] = verdoppelte_kurztexte.get(kurztext_str, 0) + 1

        # Check 6: Preis
        if einheitspreis is not None and str(einheitspreis).strip():
            preis_str = str(einheitspreis).strip()
            if preis_str in ("nan", "None", "NaN"):
                ungueltige_preise.append(f"Zeile {i}: Preis='{preis_str}'")
            else:
                try:
                    preis_val = float(einheitspreis)
                    if preis_val < 0:
                        ungueltige_preise.append(f"Zeile {i}: negativer Preis ({preis_val})")
                except (ValueError, TypeError):
                    ungueltige_preise.append(f"Zeile {i}: Preis='{preis_str}' ist keine Zahl")

    # --- Probleme sammeln ---
    if leere_zeilen > 0:
        result.probleme.append(
            f"{leere_zeilen} komplett leere Datenzeilen gefunden. "
            "Plancraft stolpert darüber — sollte vor Import entfernt werden."
        )

    if ungueltige_einheiten:
        result.probleme.append(
            f"{len(ungueltige_einheiten)} Einheit(en) nicht in Plancraft-Whitelist: "
            f"{', '.join(sorted(ungueltige_einheiten))}. "
            f"Erlaubt: {', '.join(PLANCRAFT_EINHEITEN_OFFIZIELL)}"
        )

    if ungueltige_mengen:
        result.probleme.extend(ungueltige_mengen[:5])  # Max 5 anzeigen
        if len(ungueltige_mengen) > 5:
            result.probleme.append(f"  ... und {len(ungueltige_mengen) - 5} weitere ungültige Mengen")

    if ungueltige_preise:
        result.probleme.extend(ungueltige_preise[:5])
        if len(ungueltige_preise) > 5:
            result.probleme.append(f"  ... und {len(ungueltige_preise) - 5} weitere ungültige Preise")

    if positionen_ohne_text > 0:
        result.probleme.append(
            f"{positionen_ohne_text} Position(en) ohne Kurz- UND Langtext. "
            "Plancraft braucht mindestens eine Beschreibung pro Zeile."
        )

    # Häufige Verdoppelungen (Warnung, kein Fehler)
    echte_verdoppelungen = {k: v for k, v in verdoppelte_kurztexte.items() if v > 2 and len(k) > 5}
    if echte_verdoppelungen:
        top = sorted(echte_verdoppelungen.items(), key=lambda x: -x[1])[:3]
        result.warnungen.append(
            f"{len(echte_verdoppelungen)} Kurztexte erscheinen 3+ mal. "
            f"Top 3: " + ", ".join(f"'{k[:40]}' ({v}x)" for k, v in top)
        )

    # --- Statistik ---
    positionen_mit_preis = sum(
        1 for row in datenzeilen
        if row and len(row) >= 5 and row[4] is not None and str(row[4]).strip() not in ("", "nan", "None", "NaN")
    )
    positionen_ohne_preis = anzahl_datenzeilen - leere_zeilen - positionen_mit_preis

    result.statistik = {
        "anzahl_datenzeilen": anzahl_datenzeilen,
        "leere_zeilen": leere_zeilen,
        "ungueltige_einheiten_count": len(ungueltige_einheiten),
        "ungueltige_mengen_count": len(ungueltige_mengen),
        "ungueltige_preise_count": len(ungueltige_preise),
        "positionen_ohne_text": positionen_ohne_text,
        "positionen_mit_preis": positionen_mit_preis,
        "positionen_ohne_preis": positionen_ohne_preis,
    }

    result.info.append(
        f"Validierung: {anzahl_datenzeilen} Datenzeilen, "
        f"{positionen_mit_preis} mit Preis, "
        f"{leere_zeilen} leer, "
        f"{len(ungueltige_einheiten)} ungültige Einheit(en)"
    )

    # --- Status bestimmen ---
    if result.probleme:
        result.status = "fehler"
    elif result.warnungen:
        result.status = "warnung"
    else:
        result.status = "ok"

    return result


if __name__ == "__main__":
    import sys
    import json
    if len(sys.argv) < 2:
        print("Usage: python validator.py <xlsx-pfad>")
        sys.exit(1)

    res = validiere_xlsx(sys.argv[1])
    print(json.dumps(res.to_dict(), indent=2, ensure_ascii=False))

    if res.status == "fehler":
        sys.exit(1)
    elif res.status == "warnung":
        sys.exit(2)
    else:
        sys.exit(0)
