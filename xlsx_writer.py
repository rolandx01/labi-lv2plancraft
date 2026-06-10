"""
xlsx_writer.py - Schreibt ParseResult in Plancraft-Import-XLSX

Plancraft-Import-Format (5 Spalten):
  A: Menge (Zahl)
  B: Einheit (Text: stk, m, m², psch, ...)
  C: Kurztext (Titel)
  D: Langtext (Detail-Beschreibung)
  E: Einheitspreis (optional, € pro Einheit)

Referenz: plancraft_import_dokumente.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Union
import re
from lv_parser import ParseResult, Position
from xlsx_parser import XlsxResult


# Footer-Pattern, die NICHT in Plancraft importiert werden dürfen
# (Architekten-PDFs haben oft "Seite X von Y" oder "Titelsumme: ..." mitten im Langtext)
RE_FOOTER_SEITE = re.compile(r"^\s*Seite\s+\d+\s+von\s+\d+\s*$", re.IGNORECASE)
RE_FOOTER_TITELSUMME = re.compile(r"^Titelsumme\s*:.*$", re.IGNORECASE)
RE_FOOTER_LV_INFO = re.compile(r"^\s*Leistungsverzeichnis\s*:.*$", re.IGNORECASE)
RE_FOOTER_OBJEKT = re.compile(r"^\s*Objekt\s*:.*$", re.IGNORECASE)
RE_FOOTER_ERKLAERUNG = re.compile(r"^Fachunternehmererklärung.*$", re.IGNORECASE)
RE_FOOTER_HINWEIS_LEER = re.compile(r"^Hinweis\s*:\s*$", re.IGNORECASE)


def _filtere_footer_zeilen(text: str) -> str:
    """Entfernt Footer-/Header-Zeilen aus dem Langtext, damit Plancraft
    nicht durch 'Seite 5 von 22' oder 'Titelsumme: ... EUR' verwirrt wird.

    Returns: gesäuberter Langtext.
    """
    if not text:
        return ""
    out_zeilen = []
    for zeile in text.split("\n"):
        z = zeile.strip()
        if not z:
            continue
        if (RE_FOOTER_SEITE.match(z)
                or RE_FOOTER_TITELSUMME.match(z)
                or RE_FOOTER_LV_INFO.match(z)
                or RE_FOOTER_OBJEKT.match(z)
                or RE_FOOTER_ERKLAERUNG.match(z)
                or RE_FOOTER_HINWEIS_LEER.match(z)):
            continue  # Diese Zeile überspringen
        out_zeilen.append(zeile.rstrip())
    return "\n".join(out_zeilen).strip()


# Plancraft-Spaltenbreiten (geschätzt, passen für die meisten Angebote)
SPALTEN_BREITEN = {
    "A": 12,   # Menge
    "B": 10,   # Einheit
    "C": 50,   # Kurztext
    "D": 80,   # Langtext
    "E": 14,   # Einheitspreis
}


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C5F8D", end_color="2C5F8D", fill_type="solid")


def schreibe_xlsx(result: Union[ParseResult, XlsxResult], output_pfad: str) -> dict:
    """Schreibt ParseResult in eine Plancraft-kompatible XLSX.

    Args:
        result: ParseResult vom lv_parser
        output_pfad: Zielpfad für .xlsx

    Returns:
        dict mit Statistiken (anzahl_positionen, mit_preis, ohne_preis)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Positionen"

    # Header schreiben
    # Plancraft-Originalvorlage (plancraft_import_dokumente.xlsx) verwendet EXAKT diese
    # Spaltenüberschriften. Variante wie "Einheitspreis (€)" wurde von Plancraft beim
    # Import offenbar ignoriert (alle Preise = 0,00 €). Daher: exakt diese Schreibweise.
    headers = ["Menge", "Einheit", "Kurztext", "Langtext", "Einheitspreis"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Spaltenbreiten setzen
    for col_letter, breite in SPALTEN_BREITEN.items():
        ws.column_dimensions[col_letter].width = breite

    # Header-Zeile einfrieren
    ws.freeze_panes = "A2"

    # Positionen schreiben
    row = 2
    mit_preis = 0
    ohne_preis = 0
    ungueltige_positionen = []  # Positionen, die rausgefiltert wurden (mit Begründung)

    for pos in result.positionen:
        # FIX 2: Komplett leere Positionen überspringen
        # Ein Position ist "leer" wenn:
        #   - kein Kurztext UND kein Langtext UND keine Menge
        # Plancraft stolpert über solche Zeilen und gibt "irgendein Problem" zurück.
        hat_text = (pos.kurztext and pos.kurztext.strip()) or (pos.langtext and pos.langtext.strip())
        if not hat_text and pos.menge is None:
            ungueltige_positionen.append(f"Zeile '{pos.pos_nr}': komplett leer, übersprungen")
            continue

        # FIX: Mindestens Kurztext muss da sein — sonst ist es Müll
        if not pos.kurztext or not pos.kurztext.strip():
            # Versuch, aus Langtext zu retten (erste Zeile als Kurztext)
            if pos.langtext and pos.langtext.strip():
                erste_zeile = pos.langtext.strip().split("\n")[0][:80]
                if erste_zeile.strip():
                    pos.kurztext = erste_zeile
                else:
                    ungueltige_positionen.append(f"Zeile '{pos.pos_nr}': weder Kurz- noch Langtext, übersprungen")
                    continue
            else:
                ungueltige_positionen.append(f"Zeile '{pos.pos_nr}': kein Kurztext, übersprungen")
                continue

        # Bonus-Fix: Footer-Zeilen aus dem Langtext rausfiltern
        sauberer_langtext = _filtere_footer_zeilen(pos.langtext)

        # Menge (leer lassen wenn None, statt 0)
        ws.cell(row=row, column=1, value=pos.menge if pos.menge is not None else "")

        # Einheit: Wenn leer, default "Stk." (Plancrafts Standard-Einheit)
        # Plancraft mag leere Einheiten nicht — setzt sonst auf "1,00 Stk." was oft
        # falsch ist für "m²", "l" etc. Wir geben lieber direkt Stk. mit und Labi
        # ändert es bei Bedarf in Plancraft per Klick.
        if pos.einheit:
            ws.cell(row=row, column=2, value=pos.einheit)
        else:
            ws.cell(row=row, column=2, value="Stk.")
            # FIX: Hinweis dass Einheit default-gesetzt wurde (oft bei zusammengesetzten
            # Einheiten wie StWo/mwo die Plancraft nicht direkt kennt)
            if pos.menge is not None:
                result.warnungen.append(
                    f"Zeile '{pos.pos_nr}' ({pos.kurztext[:30]}): Einheit nicht in Plancraft-Whitelist, "
                    f"default 'Stk.' gesetzt. Labi muss in Plancraft manuell korrigieren."
                )

        # Kurztext
        ws.cell(row=row, column=3, value=pos.kurztext.strip())

        # Langtext (gefiltert)
        ws.cell(row=row, column=4, value=sauberer_langtext)

        # Einheitspreis (optional)
        if pos.einheitspreis is not None:
            cell = ws.cell(row=row, column=5, value=pos.einheitspreis)
            cell.number_format = "#,##0.00 €"
            mit_preis += 1
        else:
            ws.cell(row=row, column=5, value="")
            ohne_preis += 1

        # Zeilenhöhe auto-anpassen
        ws.row_dimensions[row].height = None  # Excel macht das automatisch

        # Wrap-Text für Langtext aktivieren
        c4 = ws.cell(row=row, column=4)
        c4.alignment = Alignment(wrap_text=True, vertical="top")
        c3 = ws.cell(row=row, column=3)
        c3.alignment = Alignment(wrap_text=True, vertical="top")

        row += 1

    # FIX: Warnungen über rausgefilterte Zeilen in die Meta-Spalte
    if ungueltige_positionen:
        result.warnungen.append(
            f"{len(ungueltige_positionen)} ungültige/leere Zeile(n) automatisch entfernt. "
            "Diese Zeilen hatten weder Beschreibung noch Menge und hätten Plancrafts "
            "Import blockiert. Roh-Daten siehe Meta-Spalte."
        )

    # Zusatz-Info: In Spalte G (rechts der Positionsdaten) statt unter den Positionen.
    # Grund: Plancraft akzeptiert nur exakt 5 Spalten (Menge..Einheitspreis). Alles was nach
    # Spalte E in den Datenzeilen steht, wird ignoriert. Aber alles was UNTER den Datenzeilen
    # in Spalte A..E auftaucht, wird als neue Position interpretiert. Daher: Hinweise rechts
    # der Daten in einer "Meta"-Spalte, deutlich getrennt durch eine Leerzeile.
    if result.warnungen or result.fehler:
        # Eine zusätzliche "Meta"-Spalte in Spalte G (Index 7) — außerhalb der 5-Spalten-Spec,
        # aber Plancraft ignoriert sie. Verhindert dass Warnungen als Zeile 24/25 importiert werden.
        meta_col = 7
        meta_row = 1
        ws.cell(row=meta_row, column=meta_col, value="Hinweise").font = Font(bold=True, italic=True)
        meta_row += 1
        for warn in result.warnungen:
            ws.cell(row=meta_row, column=meta_col, value=f"⚠ {warn}").font = Font(italic=True, color="B45F06")
            meta_row += 1
        for err in result.fehler:
            ws.cell(row=meta_row, column=meta_col, value=f"✗ {err}").font = Font(italic=True, color="CC0000")
            meta_row += 1
        ws.column_dimensions[get_column_letter(meta_col)].width = 60

    wb.save(output_pfad)

    return {
        "anzahl_positionen": len(result.positionen),
        "mit_preis": mit_preis,
        "ohne_preis": ohne_preis,
        "zeilen_geschrieben": row - 2,
    }


if __name__ == "__main__":
    import sys
    from lv_parser import parse_pdf

    if len(sys.argv) < 3:
        print("Usage: python xlsx_writer.py <pdf-pfad> <xlsx-pfad>")
        sys.exit(1)

    ergebnis = parse_pdf(sys.argv[1])
    stats = schreibe_xlsx(ergebnis, sys.argv[2])
    print(f"XLSX geschrieben: {sys.argv[2]}")
    print(f"  Positionen: {stats['anzahl_positionen']}")
    print(f"  Mit Preis:  {stats['mit_preis']}")
    print(f"  Ohne Preis: {stats['ohne_preis']}")
